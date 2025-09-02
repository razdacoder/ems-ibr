from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict


class TimetableBroadSheet:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Examination Timetable"

        # Define styles
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.title_font = Font(bold=True, size=14)
        self.sub_header_font = Font(bold=True, size=10)
        self.header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid")
        self.alt_fill = PatternFill(
            start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        self.dept_fill = PatternFill(
            start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.center_alignment = Alignment(
            horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')

    def create_header_section(self, semester, academic_year, start_row=1):
        """Create the header section with title and semester info"""
        # Main title
        self.ws.merge_cells(f'A{start_row}:H{start_row}')
        title_cell = self.ws[f'A{start_row}']
        title_cell.value = f"EXAMINATION TIMETABLE - {semester} SEMESTER {academic_year}"
        title_cell.font = self.title_font
        title_cell.alignment = self.center_alignment

        # Generated date
        self.ws.merge_cells(f'A{start_row + 1}:H{start_row + 1}')
        date_cell = self.ws[f'A{start_row + 1}']
        date_cell.value = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        date_cell.alignment = self.center_alignment

        return start_row + 3  # Return next available row

    def create_column_headers(self, start_row):
        """Create column headers for the timetable"""
        headers = [
            'S/N', 'Date', 'Day', 'Period', 'Course Code',
            'Course Name', 'Class', 'Department'
        ]

        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

            # Set column widths
            column_letter = get_column_letter(col)
            if col == 1:  # S/N
                self.ws.column_dimensions[column_letter].width = 6
            elif col == 2:  # Date
                self.ws.column_dimensions[column_letter].width = 12
            elif col == 3:  # Day
                self.ws.column_dimensions[column_letter].width = 12
            elif col == 4:  # Period
                self.ws.column_dimensions[column_letter].width = 15
            elif col == 5:  # Course Code
                self.ws.column_dimensions[column_letter].width = 15
            elif col == 6:  # Course Name
                self.ws.column_dimensions[column_letter].width = 40
            elif col == 7:  # Class
                self.ws.column_dimensions[column_letter].width = 20
            elif col == 8:  # Department
                self.ws.column_dimensions[column_letter].width = 25

        return start_row + 1  # Return next available row

    def populate_timetable_data(self, timetables, start_row):
        """Populate the timetable data grouped by department"""
        current_row = start_row
        serial_number = 1

        # Group timetables by department for better organization
        dept_groups = defaultdict(list)
        for timetable in timetables:
            dept_groups[timetable.class_obj.department.name].append(timetable)

        for dept_name, dept_timetables in dept_groups.items():
            # Add department header
            self.ws.merge_cells(f'A{current_row}:H{current_row}')
            dept_cell = self.ws[f'A{current_row}']
            dept_cell.value = f"DEPARTMENT: {dept_name.upper()}"
            dept_cell.font = self.sub_header_font
            dept_cell.fill = self.dept_fill
            dept_cell.alignment = self.center_alignment
            dept_cell.border = self.thin_border
            current_row += 1

            # Sort department timetables by date, then by period
            dept_timetables.sort(key=lambda x: (x.date, x.period))

            for timetable in dept_timetables:
                # Apply alternating row colors
                fill_color = self.alt_fill if serial_number % 2 == 0 else None

                row_data = [
                    serial_number,  # S/N
                    timetable.date.strftime('%d/%m/%Y'),  # Date
                    timetable.date.strftime('%A'),  # Day
                    timetable.period,  # Period
                    timetable.course.code,  # Course Code
                    timetable.course.name,  # Course Name
                    timetable.class_obj.name,  # Class
                    timetable.class_obj.department.name  # Department
                ]

                for col, value in enumerate(row_data, 1):
                    cell = self.ws.cell(row=current_row, column=col)
                    cell.value = value
                    cell.border = self.thin_border

                    if col == 1 or col == 4:  # S/N and Period columns
                        cell.alignment = self.center_alignment
                    else:
                        cell.alignment = self.left_alignment

                    if fill_color:
                        cell.fill = fill_color

                current_row += 1
                serial_number += 1

            # Add space between departments
            current_row += 1

        return current_row

    def add_summary_section(self, timetables, start_row):
        """Add summary statistics"""
        current_row = start_row + 1

        # Summary title
        self.ws.merge_cells(f'A{current_row}:H{current_row}')
        summary_cell = self.ws[f'A{current_row}']
        summary_cell.value = "EXAMINATION SUMMARY"
        summary_cell.font = self.sub_header_font
        summary_cell.alignment = self.center_alignment
        current_row += 2

        # Statistics
        total_exams = len(timetables)
        total_courses = len(set(t.course.id for t in timetables))
        total_departments = len(
            set(t.class_obj.department.id for t in timetables))

        stats = [
            f"Total Examinations: {total_exams}",
            f"Total Courses: {total_courses}",
            f"Total Departments: {total_departments}",
            f"Examination Period: {min(t.date for t in timetables).strftime('%d/%m/%Y') if timetables else 'N/A'} - {max(t.date for t in timetables).strftime('%d/%m/%Y') if timetables else 'N/A'}"
        ]

        for stat in stats:
            cell = self.ws.cell(row=current_row, column=1)
            cell.value = stat
            cell.font = Font(size=10)
            self.ws.merge_cells(f'A{current_row}:H{current_row}')
            current_row += 1

    def add_footer_notes(self, start_row):
        """Add footer notes and instructions"""
        notes = [
            "EXAMINATION INSTRUCTIONS:",
            "1. Students must arrive at the examination venue 15 minutes before the scheduled time",
            "2. No student will be allowed into the examination hall 30 minutes after the exam has commenced",
            "3. Students must bring valid student ID cards and required stationery",
            "4. Mobile phones and electronic devices are strictly prohibited in examination halls",
            "5. Any form of examination malpractice will result in disqualification",
            "6. Students should check the examination venue and time carefully"
        ]

        current_row = start_row + 2

        for note in notes:
            cell = self.ws.cell(row=current_row, column=1)
            if note.startswith("EXAMINATION INSTRUCTIONS:"):
                cell.font = Font(bold=True, size=11)
            else:
                cell.font = Font(size=10)

            self.ws.merge_cells(f'A{current_row}:H{current_row}')
            cell.value = note
            cell.alignment = self.left_alignment
            current_row += 1

    def generate_excel(self, timetables, semester="2ND", academic_year="2024/2025"):
        """Generate the complete Excel file"""
        # Create header section
        current_row = self.create_header_section(semester, academic_year)

        # Create column headers
        current_row = self.create_column_headers(current_row)

        # Populate data
        current_row = self.populate_timetable_data(timetables, current_row)

        # Add summary section
        self.add_summary_section(timetables, current_row)
        current_row += 8  # Space for summary

        # Add footer notes
        self.add_footer_notes(current_row)

        return self.wb
